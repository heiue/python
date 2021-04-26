#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time : 2021/4/25 2:16 下午
# @Author : weiyizheng
# @File : makeExport.py
# @desc :
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import fonts, Alignment, colors
import time
import os
import sys
import datetime
from dateutil.relativedelta import relativedelta


class MakeExport(object):
    list_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'G', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                   'U',
                   'V', 'W', 'X', 'Y', 'Z']

    topTitle = [
        {
            "name": "年限"
        },
        {
            "name": "付款方式"
        },
        {
            "name": "租金明细",
            "needCell": 5
        }
    ]

    execlName = ''

    def __init__(self, name=None, sheetName=None, total_pay:int=None, total_month:int=None, how_month_pay:int=None,
                 start_date=None):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        if sheetName is None:
            self.ws1.title = 'sheet1'
        else:
            self.ws1.title = sheetName
        if name is None:
            self.execlName = os.getcwd() + '/../file/' + time.strftime("%Y-%m-%d", time.localtime()) + '.xlsx'
        else:
            self.execlName = os.getcwd() + '/../file/' + name + '.xlsx'

        if total_pay is None:
            raise Exception('总款为空')
        else:
            self.total_pay = total_pay
        if total_month is None:
            raise Exception('总月为空')
        else:
            self.total_month = total_month
        if how_month_pay is None:
            raise Exception('付款间隔为空')
        else:
            self.how_month_pay = how_month_pay
        if start_date is None:
            raise Exception('签约开始日期为空')
        else:
            self.start_date = start_date

        # 设置全部水平居中
        for item in self.list_letter:
            self.ws1[item + '1'].alignment = Alignment(horizontal='center', vertical='center')

    def top_header(self):
        point = 1
        for item in self.topTitle:
            self.ws1.cell(column=point, row=1, value=item['name'])
            if 'needCell' in item.keys():
                start_col = i = point
                end_col = point + int(item['needCell'])
                while i <= end_col:
                    column_letter = self.num_to_letter(i)
                    self.ws1.column_dimensions[column_letter].width = 30
                    i += 1

                self.ws1.merge_cells(start_row=1, start_column=point, end_row=1, end_column=end_col)
                point += item['needCell']
            else:
                point += 1

    def data_list(self):
        # 每月要付多少钱
        month_pay = int(self.total_pay / self.total_month)
        # 要付多少期
        stage_pay = int(self.total_month / self.how_month_pay)

        # 合并第一列
        self.ws1.cell(row=2, column=1, value=(self.total_month/12))
        self.ws1.merge_cells(start_row=2, start_column=1, end_row=stage_pay+2, end_column=1)
        # 合并第二列
        self.ws1.cell(row=2, column=2, value=self.how_month_pay)
        self.ws1.merge_cells(start_row=2, start_column=2, end_row=stage_pay+2, end_column=2)

        self.start_date = self.date_to_datetime()

        for i in range(3, stage_pay + 2):
            end_date = self.start_date + relativedelta(months=+self.how_month_pay) - relativedelta(days=+1)
            value = self.start_date.strftime("%Y-%m-%d") + '至' + end_date.strftime("%Y-%m-%d")
            self.ws1.cell(column=3, row=i, value=value)
            self.start_date += relativedelta(months=+self.how_month_pay)

    def run(self):
        self.top_header()
        self.data_list()
        self.wb.save(filename=self.execlName)

    def num_to_letter(self, num):
        return self.list_letter[num - 1]

    def date_to_datetime(self):
        time_array = time.strptime(self.start_date, "%Y-%m-%d")
        return datetime.date(time_array.tm_year, time_array.tm_mon, time_array.tm_mday)


# now_time = datetime.datetime.now()
# n_time = datetime.date(2021, 1, 1) + relativedelta(months=+3) - relativedelta(days=+1)
# n_time = time.strptime('2021-01-01', "%Y-%m-%d")
# print(n_time)

# exit()

params = sys.argv
print(params)
try:
    # 总款
    total_pay = int(params[1])
    # 总月
    total_month = int(params[2])
    # 付款间隔
    how_month_pay = int(params[3])
    if int(total_month) % int(how_month_pay) != 0:
        raise Exception('不能整除')
    # 签约开始日期
    start_date = params[4]
except IndexError as e:
    print(e.args[0])
    exit()
except Exception as e:
    print(e.args[0])
    exit()

MakeExport(total_pay=total_pay, total_month=total_month, how_month_pay=how_month_pay, start_date=start_date).run()
