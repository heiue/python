from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'file/empty_book2.xlsx'
ws1 = wb.active  # 第一个表
ws1.title = "sheet1"  # 第一个表命名
# 遍历第一个表的1到40行，赋值一个600内的随机数
for row in range(1, 5):
    for col in range(1, 3):
        cell = get_column_letter(col) + str(row)
        ws1[cell] = col

# 合并单元格
ws1.merge_cells('A6:A7')
ws1['A6'] = 'abc'
# ws2 = wb.create_sheet(title="Pi")
# ws2['F5'] = 3.14123123123
# ws3 = wb.create_sheet(title="Data")
# for row in range(1, 20):
#     for col in range(1, 40):
#         _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
wb.save(filename=dest_filename)
